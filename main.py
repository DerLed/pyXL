from tkinter import *
import string
import certificate
from openpyxl import load_workbook


class DbStroke:
    __column_template = [2, 4, 5, 6, 7, 8]
    row_data = []

    def __init__(self, path_book, row=5):
        my_workbook = load_workbook(path_book)
        my_worksheet = my_workbook.active
        if my_worksheet['A' + str(row)].value:
            self.row_data = [x.value for x in my_worksheet[row]]

            for x in self.__column_template:
                self.row_data[x] = None

            bt = []
            for old_value in self.row_data:
                if old_value:
                    bt.append(old_value)
            self.row_data = bt

            buffer_list = self.row_data[18:33]
            del self.row_data[18:33]
            self.row_data.append(buffer_list)

        else:
            self.row_data = ['В данной строке нет значений']

    def print(self):
        print(self.row_data)


root = Tk()
root.title('openXL')

l1 = 'C:/Копия журнал верификации лист 09Г2С - копия.xlsx'
# print(my_worksheet.max_row)

thickness = 10
count = 0
# column_template = [2, 4, 5, 6, 7, 8]
my_list = []
my_db = {}

# for my_row in range(5, my_worksheet.max_row):
#     if not my_worksheet['A' + str(my_row)].value:
#         break
#     me_row = my_worksheet[my_row]
#     super_list = [x.value for x in me_row]
#     my_list.append(super_list)

db2 = DbStroke(l1, 5)
db2.print()


count = 0
#for x in my_list[0]:
#    # print('[', count, ']', x)
 #   count += 1

h = 'Лист г/к  8*1500*6000'
y = h.split('Лист г/к')
h = h[8:-10].strip()

# print(h)
# print(my_worksheet.cell(row = 6, column = column_s).value, end=' ---  ')
# for x in range(6, my_worksheet.max_row):
#     if not my_worksheet['M' + str(x)].value:
#         break
#     print(my_worksheet['M' + str(x)].value.split())


# root.mainloop()
# bob = certificate.Certificate(my_list[0])
# print('************', bob.thickness)
# print(bob.limit_fluidity)
# print('************' * 5)
# db1 = DbStroke(my_list[0])


"""
{
  "melting_number": "?",
  "certificate_number": "?",
  "date_of_certificate": "?",
  "metal_grade": "?",
  "thickness": "?"
}
"""
